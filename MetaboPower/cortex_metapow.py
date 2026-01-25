# ==============================================================================
# Copyright (c) 2026 Andrea Bonvicin - bFactor Project
# PROPRIETARY LICENSE - TUTTI I DIRITTI RISERVATI
# Sharing, distribution or reproduction is strictly prohibited.
# La condivisione, distribuzione o riproduzione è severamente vietata.
# ==============================================================================

"""
CORTEX_METAPOW.PY - Parser per Cortex XLSX
Gestisce l'importazione di file XLSX da Cortex Metabolic System
Estrae: intestazioni (riga 117), dati (da riga 119), metadati soggetto (celle specifiche)
"""

import logging
import pandas as pd
import openpyxl
from typing import Dict, Optional
from .data_extraction_metapow import find_column, normalize_decimals

logger = logging.getLogger(__name__)


class CortexMetabolitParser:
    """Parser specializzato per file XLSX da Cortex"""
    
    def __init__(self):
        self.data = None
        self.metadata = {}
        self.file_path = None
    
    def load(self, file_path: str, settings: Dict) -> Dict:
        """Carica file XLSX da Cortex usando il profilo fornito"""
        try:
            self.file_path = file_path
            self.metadata = {}
            
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Leggi le configurazioni dal profilo
            header_row = settings.get("header_row", 117)
            subheader_row = settings.get("subheader_row", 118)
            data_start_row = settings.get("data_start_row", 119)
            metadata_cells = settings.get("metadata_cells", {})
            
            # ===== ESTRAI METADATI =====
            self._extract_metadata(ws, metadata_cells)
            
            # ===== ESTRAI INTESTAZIONI =====
            columns = self._extract_columns(ws, header_row, subheader_row)
            
            if not columns:
                return {
                    "success": False,
                    "error": f"Nessuna colonna trovata alla riga {header_row}"
                }
            
            # ===== ESTRAI DATI =====
            data_rows = self._extract_data_rows(ws, data_start_row, len(columns))
            
            if not data_rows:
                return {
                    "success": False,
                    "error": f"Nessun dato trovato da riga {data_start_row}"
                }
            
            # Crea DataFrame
            df = pd.DataFrame(data_rows, columns=columns)
            df = normalize_decimals(df)
            
            # ===== POSTPROCESSING METADATI =====
            self._postprocess_metadata(df)
            
            self.data = df
            
            return {
                "success": True,
                "file_type": "XLSX (Cortex)",
                "rows": len(df),
                "columns": columns,
                "metadata": self.metadata
            }
        
        except Exception as e:
            return {
                "success": False,
                "error": f"Errore parsing Cortex XLSX: {str(e)}"
            }
    
    def _extract_metadata(self, ws, metadata_cells: Dict) -> None:
        """Estrae metadati soggetto dalle celle specificate (es. C27, C28, C110)"""
        for key, cell_ref in metadata_cells.items():
            try:
                value = ws[cell_ref].value
                self.metadata[key] = value
            except Exception:
                self.metadata[key] = None
    
    def _extract_columns(self, ws, header_row: int, subheader_row: Optional[int]) -> list:
        """Estrae nomi colonne da header_row e aggiunge unità da subheader_row"""
        columns = []
        
        for col_idx in range(1, 100):  # Cerca fino a 100 colonne
            cell_value = ws.cell(row=header_row, column=col_idx).value
            if cell_value is None:
                break
            
            col_name = str(cell_value).strip()
            
            # Aggiungi unità se esiste subheader
            if subheader_row:
                unit = ws.cell(row=subheader_row, column=col_idx).value
                if unit and str(unit).strip():
                    col_name = f"{col_name} ({unit})"
            
            columns.append(col_name)
        
        return columns
    
    def _extract_data_rows(self, ws, data_start_row: int, num_columns: int) -> list:
        """Estrae righe dati da data_start_row fino alla prima riga vuota"""
        data_rows = []
        row_idx = data_start_row
        
        while True:
            # Controlla se la prima colonna è vuota (fine dati)
            first_cell = ws.cell(row=row_idx, column=1).value
            if first_cell is None:
                break
            
            # Leggi la riga
            row_data = []
            for col_idx in range(1, num_columns + 1):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(cell_val)
            
            data_rows.append(row_data)
            row_idx += 1
        
        return data_rows
    
    def _postprocess_metadata(self, df: pd.DataFrame) -> None:
        """Arricchisce metadati con info utili estratte dai dati:
        - turn_index: indice della riga con "Turn" nel colonna Marker (colonna 3)
        - time_column: nome della colonna tempo
        - power_column: nome della colonna potenza
        - ramp_start_index: indice della riga dove inizia la rampa (colonna W passa da 0 a >0)
        """
        # Turn marker: cerca "Turn" nella colonna 3 (Marker)
        if len(df.columns) > 2:
            marker_col = df.iloc[:, 2]  # Colonna 3 (indice 2)
            for idx, val in enumerate(marker_col):
                if val and isinstance(val, str) and val.strip().lower() == "turn":
                    self.metadata["turn_index"] = idx
                    break
        
        # Identifica colonne tempo e potenza
        time_col = find_column(df, keywords=["time", "tempo", "t ("])
        if time_col:
            self.metadata["time_column"] = time_col
        
        power_col = find_column(df, keywords=["wr", "watt", "potenza", "power"])
        if power_col:
            self.metadata["power_column"] = power_col
        
        # Rileva inizio rampa: colonna W passa da 0 a >0
        try:
            col_w = None
            
            # Prova prima a trovare colonna W per nome
            col_w = find_column(df, keywords=["W", "wr", "watt"])
            
            # Se non trovata per nome, prova indice 22 (23esima colonna)
            if not col_w and len(df.columns) > 22:
                col_w_by_idx = df.iloc[:, 22]
                # Controlla se ha valori numerici (è effettivamente una colonna W)
                try:
                    pd.to_numeric(col_w_by_idx, errors='raise')
                    col_w = df.columns[22]
                except (ValueError, TypeError):
                    pass
            
            if not col_w:
                # Nessuna colonna W trovata
                self.metadata["ramp_start_index"] = 0
                return
            
            # Estrai colonna e converti a numerico
            col_data = df[col_w]
            col_numeric = pd.to_numeric(col_data, errors='coerce')
            
            # Cerca la transizione da <=0 a >0
            # Tolleranza: considera come "start" dove power >5W
            threshold = 5.0
            for idx in range(len(col_numeric)):
                val = col_numeric.iloc[idx]
                if pd.notna(val) and val > threshold:
                    self.metadata["ramp_start_index"] = idx
                    return
            
            # Se non trova transizione, assume index 0
            self.metadata["ramp_start_index"] = 0
        
        except Exception as e:
            logger.error(f"Errore rilevamento inizio rampa: {e}", exc_info=True)
            self.metadata["ramp_start_index"] = 0
    
    
    def get_data(self) -> Optional[pd.DataFrame]:
        """Ritorna il DataFrame caricato"""
        return self.data
    
    def get_metadata(self) -> Dict:
        """Ritorna i metadati estratti"""
        return self.metadata
